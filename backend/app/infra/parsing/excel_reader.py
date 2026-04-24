from __future__ import annotations

from io import BytesIO
from typing import Any, Literal

from openpyxl import load_workbook

from .base import BaseParser
from .pymupdf_reader import ParsedDocument


def _row_to_text(headers: list[str], values: list[object]) -> str:
    parts: list[str] = []
    for i, v in enumerate(values):
        if v is None:
            continue
        key = headers[i] if i < len(headers) and headers[i] else f"col_{i+1}"
        parts.append(f"{key}: {v}")
    return "; ".join(parts).strip()


class ExcelReader(BaseParser):
    def parse(self, data: bytes, params: dict[str, Any]) -> ParsedDocument:
        mode: Literal["row_to_text", "chunk_by_rows"] = params.get("mode", "row_to_text")
        rows_per_chunk = int(params.get("rows_per_chunk", 50))
        if rows_per_chunk <= 0:
            raise RuntimeError("rows_per_chunk must be > 0")

        try:
            wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        except Exception as e:
            raise RuntimeError(
                "Failed to parse Excel file. Only .xlsx is supported reliably (please convert .xls → .xlsx)."
            ) from e
        try:
            parts: list[str] = []
            for ws in wb.worksheets:
                rows_iter = ws.iter_rows(values_only=True)
                try:
                    header_row = next(rows_iter)
                except StopIteration:
                    continue
                headers = [("" if c is None else str(c)).strip() for c in header_row]

                buffer: list[str] = []
                for row in rows_iter:
                    text = _row_to_text(headers, list(row))
                    if not text:
                        continue
                    if mode == "row_to_text":
                        parts.append(text)
                    elif mode == "chunk_by_rows":
                        buffer.append(text)
                        if len(buffer) >= rows_per_chunk:
                            parts.append("\n".join(buffer))
                            buffer = []
                    else:
                        raise RuntimeError(f"Unsupported excel_reader.mode: {mode}")

                if mode == "chunk_by_rows" and buffer:
                    parts.append("\n".join(buffer))
        finally:
            wb.close()

        text = "\n\n".join(parts).strip()
        if not text:
            raise RuntimeError("Parsed document is empty")
        return ParsedDocument(text=text, page_count=1)

